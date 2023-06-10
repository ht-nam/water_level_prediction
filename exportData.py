import numpy as np
import matplotlib.pyplot as plt

plt.switch_backend("agg")
import openpyxl

from measure import NSE, R2, MAE, RMSE, MAX_ERROR


def nse(predictions, targets):
    return 1 - (
        np.sum((predictions - targets) ** 2) / np.sum((targets - np.mean(targets)) ** 2)
    )


def getResult(
    x_test,
    y_test,
    y_prd,
    callbackDays,
    threshold,
    cols,
    labelCol,
    folderName,
    refValue,
    reference_col,
):
    temp1 = 0
    for i in range(len(y_test)):
        if np.abs(y_test[i] - y_prd[i]) > threshold:
            temp1 += 1

    x_test = x_test.reshape(x_test.shape[0], -1)

    # export excel result
    data_cols = []
    # data_cols = cols * callbackDays
    for i in range(1, callbackDays + 1):
        for j in cols:
            data_cols.append(j + "_" + str(i))
    if reference_col == "":
        input_detail = [data_cols + ["REAL_" + labelCol, "PRD_" + labelCol, "ABS"]]
        for i in range(0, x_test.shape[0]):
            input_detail.append(
                x_test[i].tolist()
                + y_test[i].tolist()
                + y_prd[i].tolist()
                + [abs(y_prd[i][0] - y_test[i][0])]
            )
    else:
        input_detail = [
            data_cols + ["REAL_" + labelCol, "PRD_" + labelCol, "ABS", reference_col]
        ]
        for i in range(0, x_test.shape[0]):
            input_detail.append(
                x_test[i].tolist()
                + y_test[i].tolist()
                + y_prd[i].tolist()
                + [abs(y_prd[i][0] - y_test[i][0])]
                + [refValue[i]]
            )
    filename = callbackDays
    output_excel_path = folderName + "/callbackDay" + str(filename) + ".xlsx"
    measure = [
        ["OTR", temp1 / y_test.shape[0]],
        ["Max error", MAX_ERROR(y_test, y_prd)[0]],
        ["R2 score", R2(y_test, y_prd)],
        ["NSE score", NSE(y_test, y_prd)],
        ["MAE score", MAE(y_test, y_prd)],
        ["RMSE score", RMSE(y_test, y_prd)],
    ]
    output_Excel(input_detail, output_excel_path, measure)

    plotResult(y_test, y_prd, folderName + "/callbackDay" + str(filename))
    return [
        callbackDays,
        y_test.shape[0],
        temp1,
        temp1 / y_test.shape[0],
        MAX_ERROR(y_test, y_prd)[0],
        R2(y_test, y_prd),
        NSE(y_test, y_prd),
        MAE(y_test, y_prd),
        RMSE(y_test, y_prd),
    ]


def plotResult(y_test, y_prd, imglnk):
    plt.plot(y_test)
    plt.plot(y_prd)
    plt.savefig(imglnk)
    plt.cla()


def output_Excel(input_detail, output_excel_path, measure=[]):
    row = len(input_detail)
    column = len(input_detail[0])

    wb = openpyxl.Workbook()
    ws = wb.active

    for i in range(0, row):
        for j in range(0, column):
            v = input_detail[i][j]
            ws.cell(column=j + 1, row=i + 1, value=v)

    if len(measure) > 0:
        for i in range(0, len(measure)):
            for j in range(0, len(measure[0])):
                v = measure[i][j]
                ws.cell(column=j + 1, row=i + 1 + row + 1, value=v)

    wb.save(output_excel_path)
