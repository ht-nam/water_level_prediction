from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout, LSTM
import os
from getData import loadData, y_scaler
from exportData import printResult
from openpyxl import Workbook
import time


trainFile = "dataset\Train_data_WL_RF_21_22.csv"
testFile = "dataset\Test_data_WL_RF_21_22.csv"
knowCols = [
    "WL_KienGiang",
    "RF_KienGiang",
    "WL_LeThuy",
    "RF_LeThuy",
    "WL_DongHoi",
    "RF_DongHoi",
    "Tide_DongHoi",
]
labelCol = "WL_LeThuy"
callbackTime = 6
stepTime = 6
epochs = 10
batchSize = 30
waterLevel = 1.2
isSmote = True


def lstmModel(
    trainFile,
    testFile,
    knowCols,
    labelCol,
    foldername,
    know_attributes,
    callbackTime=6,
    stepTime=6,
    epochs=100,
    batchSize=30,
    waterLevel=1,
    isSmote=False,
):
    if not os.path.exists(foldername):
        os.makedirs(foldername)

    # foldername_split = foldername.split("/")
    # folderchild = foldername_split[-1]
    # Sumary_result_file = foldername + '/Sumary_result_' + folderchild + '.xlsx'
    # feature_attributes = know_attributes + knowCols
    # attributes=[]
    # for i in range(len(feature_attributes)):
    #     if feature_attributes[i] not in labelCol:
    #         attributes.append(feature_attributes[i])
    # for i in range(len(feature_attributes)):
    #     if feature_attributes[i] in labelCol:
    #         attributes.append(feature_attributes[i])
    # print('attributes:', attributes)
    # all_attributes = []
    # for i in range(len(attributes)):
    #     if attributes[i] not in labelCol:
    #         all_attributes.append(attributes[i])
    # for i in range(len(labelCol)):
    #     all_attributes.append(labelCol[i])
    # print('all_attributes:', all_attributes)

    # wb = Workbook()
    # ws = wb.active
    # ws.cell(column=1, row=1, value='Afterdays')
    # ws.cell(column=2, row=1, value='Numdays')
    # ws.cell(column=3, row=1, value='NSE')
    # ws.cell(column=4, row=1, value='R2')
    # ws.cell(column=5, row=1, value='MAE')
    # ws.cell(column=6, row=1, value='RMSE')
    # ws.cell(column=7, row=1, value='OTR')
    # ws.cell(column=8, row=1, value='MAX error')
    # ws.cell(column=9, row=1, value='Time (sec)')

    # if (len(know_attributes)==0):
    #     start_index = 2
    #     # callbackTime += 1
    # else:
    #     start_index = 1

    # normalize = False
    # for afterdays in range(callbackTime, callbackTime + 1):
    #     for numdays in range(start_index, stepTime):
    #         count += 1
    #         numdays1 = numdays-(start_index-1)
    #         filename = foldername + '/result_' + folderchild + '_afterdays_' + str(afterdays) + '_numdays_'+ str(numdays1)
    #         output_file = filename + '.xlsx'
    #         start_time = time.time()
    #         X_test_original, y_test_original, X_train, y_train, X_test, y_test, scalerX, scalerY = loadData(trainFile=trainFile,
    #                                                                                                         testFile=testFile,
    #                                                                                                         cols=knowCols,
    #                                                                                                         label_col=labelCol,
    #                                                                                                         step_days=stepTime,
    #                                                                                                         callback_days=callbackTime,
    #                                                                                                         water_level=waterLevel,
    #                                                                                                         is_smote=isSmote)



    # get data
    x_train, y_train, x_test, y_test = loadData(
        trainFile=trainFile,
        testFile=testFile,
        cols=knowCols,
        label_col=labelCol,
        step_days=stepTime,
        callback_days=callbackTime,
        water_level=waterLevel,
        is_smote=isSmote,
    )

    # train model
    model = Sequential()
    model.add(
        LSTM(
            128,
            activation="tanh",
            input_shape=(x_train.shape[1], x_train.shape[2]),
            return_sequences=False,
        )
    )
    model.add(Dense(y_train.shape[1], activation="tanh")),

    model.compile(optimizer="adam", loss="mse")
    model.summary()
    model.fit(
        x_train,
        y_train,
        epochs=epochs,
        batch_size=batchSize,
        validation_split=0.3,
        verbose=1,
    )
    # end train model

    # predict
    y_prd = model.predict(x_test)
    y_test_inverse = y_scaler.inverse_transform(y_test)
    y_prd_inverse = y_scaler.inverse_transform(y_prd)
    printResult(y_test_inverse, y_prd_inverse, [], callbackTime)


# lstmModel(
#     trainFile,
#     testFile,
#     knowCols,
#     labelCol,
#     callbackTime,
#     stepTime,
#     epochs,
#     batchSize,
#     waterLevel,
#     isSmote,
# )
